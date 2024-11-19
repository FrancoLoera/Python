import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment
import random
import datetime

libro = openpyxl.Workbook()
hoja = libro["Sheet"]

#Generar listado de números y sumarlos
hoja["B1"].value = "Listado de números"

for numero in range(2, 12):
    hoja.cell(row = numero, column = 1).value = random.randint(0, 100)
    
hoja["B12"].value = "=SUM(A2:A12)"

#Presentar fecha actual
hoja["D1"].value = "Hoy es:"
hoja["E1"].value = datetime.datetime.now().strftime("%d/%m/%Y")

#Guardar el libro
libro.save("ExcelPython.xlsx")

#Formato de celdas
#Letra bold
celdas_negritas = hoja["B1"]
negritas = Font(bold = True)
celdas_negritas.font = negritas

#Ancho de columna
hoja.column_dimensions["B"].width = len(celdas_negritas.value)

#Borde inferior
border_inferior = Border(bottom=Side(border_style="thin", color="000000"))
celdas_negritas.border = border_inferior

#Centrando contenido
centrado = Alignment(horizontal="center", vertical="center")

for renglon in range(2, 12):
    hoja.cell(row = renglon, column = 1).alignment = centrado

libro.save("ExcelPython.xlsx")

#Lectura
libro = openpyxl.load_workbook("ExcelPython.xlsx")
print(type(libro))
print(libro.sheetnames)

hoja = libro["Sheet"]

#Acceso a datos mediante key
print(hoja["A1"].value)
print(hoja["A2"].value)

#Acceso a datos mediante coordenada
coordenada = hoja.cell(row=3, column=1).coordinate
print('Coordenada: ', coordenada)

for fila in range(2, 12):
    print(hoja.cell(row=fila, column=1).value)

#Método Pythónico
rango_celdas = hoja["A2":"A12"]
print(rango_celdas)

for renglon in rango_celdas:
    for celda in renglon:
        print(f"Celda: {celda.coordinate} = {celda.value}")