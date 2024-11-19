import openpyxl
from openpyxl.styles import Font, Side, Border, Alignment

#Creación del libro
libro = openpyxl.Workbook()

#Creación de la hoja
hoja = libro["Sheet"]
hoja.title = "Datos de Empleados"

#Contenido y formato de Encabezados de columna
hoja["A1"].value = "ID"
hoja["B1"].value = "Nombre/s"
hoja["C1"].value = "Apellido/s"

#Borde total para Encabezados
for i in range(1, 4):
	hoja.cell(row = 1, column = i).border = Border(bottom = Side(border_style = "medium", color = "000000"), right = Side(border_style = "medium", color = "000000"), left = Side(border_style = "medium", color = "000000"), top = Side(border_style = "medium", color="000000"))
	
	#Negritas para Encabezados
	hoja.cell(row = 1, column = i).font = Font(bold = True)

#Ancho de columnas
hoja.column_dimensions["A"].width = 10
hoja.column_dimensions["B"].width = 20
hoja.column_dimensions["C"].width = 20

#Alineación de encabezado ID
alineacion_derecha = Alignment(horizontal = "right", vertical = "center")

hoja["A1"].alignment = alineacion_derecha

dict_Datos = dict()

#Cíclo de datos
while True:
	try:
		nombres = input("Ingrese los nombres del empleado: ")
	
		#Validación anti-nulos
		if nombres.replace(" ","") == "" or nombres.isnumeric(): break
	
		apellidos = input("Ingrese los apellidos del empleado: ")
	
		#Validación anti-nulos
		if apellidos.replace(" ","") == "" or apellidos.isnumeric(): break
	
	except Exception:
		print("Ups... Algo ha salido mal\n")
	
	else:			
		nombres = nombres.capitalize()
		apellidos = apellidos.capitalize()
	
		#Asignación de ID
		dict_Datos[max(dict_Datos, default=0) + 1] = (nombres, apellidos)
		type(dict_Datos)
	
		opc = input("Introduzca 'X' si desea terminar el proceso: ")
		print("\n")
	
		if opc.upper() == "X": break

i = 2
for clave in dict_Datos.keys():
    
	hoja.cell(row = i, column = 1).value = clave
	hoja.cell(row = i, column = 2).value = dict_Datos[clave][0]
	hoja.cell(row = i, column = 3).value = dict_Datos[clave][1]

	if i < len(dict_Datos) + 1: hoja.cell(row = i, column = 3).border = Border(right = Side(border_style = "medium", color="000000"))
	
	i += 1

	if i == len(dict_Datos) + 1:
		hoja.cell(row = i, column = 1).border = Border(bottom = Side(border_style="medium", color="000000"))
		hoja.cell(row = i, column = 2).border = Border(bottom = Side(border_style="medium", color="000000"))
		hoja.cell(row = i, column = 3).border = Border(bottom = Side(border_style="medium", color="000000"), right = Side(border_style = "medium", color="000000"))
		
libro.save("Libro_Empleados.xlsx")